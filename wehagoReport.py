from typing import Tuple
from openpyxl.xml.constants import SHEET_DRAWING_NS
import pandas as pd  
import numpy as np  
import glob
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image
import sys, os, datetime, time
import xlsxwriter
import matplotlib as mpl
import matplotlib.pyplot as plt
import varname
from wehagotest import Login, browser_click
import platform
from selenium import webdriver
from driver import hasxpath, currentTime

def resultPath() :
    path = os.path.join(os.getcwd(), 'result')
    return path

def detailPath() :
    path = os.path.join(os.getcwd(), 'result')
    path = os.path.join(path, 'detail')
    return path

def today(date=None) :
    day = datetime.datetime.now()
    if date :
        day = day + datetime.timedelta(days=date)
        day = day.strftime('%Y-%m-%d')
    else :
        day = day.strftime('%Y-%m-%d %Hh%Mm')
    return day

# 테스트 결과 리포트
def testReport(version) :
    now = today()
    if version == 2:
        file_name = '/' + now + ' WEHAGO TEST 결과_업데이트.xlsx'
    elif version == 3 :
        file_name = '/' + now + ' WEHAGO TEST 결과_발행.xlsx'
    else :
        file_name = '/' + now + ' WEHAGO TEST 결과.xlsx'
    return file_name

# glob 1시간 파일 리포트
def hourReport(version) : 
    now = today()[:14]
    if version == 2:
        file_name = detailPath() + '/*' + now + '* TEST 결과_업데이트.xlsx'
    elif version == 3:
        file_name = detailPath() + '/*' + now + '* TEST 결과_발행.xlsx'
    else :
        file_name = detailPath() + '/*' + now + '* TEST 결과.xlsx'
    return file_name

# glob 24시간 파일 리포트
def dayReport(version) :
    now = today(-1)
    if version == 2:
        now = today()[:14]
        file_name = detailPath() + '/*' + now + '* TEST 결과_업데이트.xlsx'
    elif version == 3:
        file_name = detailPath() + '/*' + now + '* TEST 결과_발행.xlsx'
    else :
        file_name = detailPath() + '/*' + now + '* TEST 결과.xlsx'
    return file_name

# 업로드 리포트
def wehagoReport(version, brand, hour=True) :
    if hour :
        now = today()[:14]
    else :
        now = today(-1)

    if version == 2 :
        file_name = resultPath() + '/' + now[:10] + '_' + brand + '업데이트.xlsx'
    elif version == 3 :
        file_name = resultPath() + '/' + now + '_' + brand + '발행.xlsx'
    else :
        file_name = resultPath() + '/' + now + '_' + brand + '통합.xlsx'

    return file_name

def reportRun (version, brand, hour=True) :
    if brand == 2 :
        brand = 'T_'
    elif brand == 3 :
        brand = 'V_'
    else : 
        brand = 'W_'
    WehagoResult().fileUnion(version, brand, hour)
    if not isNullFile(version, brand, hour) :
        path = os.getcwd()
        if platform.system() == 'Windows' :
            browser = webdriver.Chrome(path + '\\chromedriver.exe')
        elif platform.system() == 'Darwin' :
            browser = webdriver.Chrome(path + '/chromedriver')
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("detach", True)
        # 사이즈 조정 
        browser.maximize_window()
        Login().login(browser, 'hancho01')
        WehagoResult().upload(browser, version, brand, hour)
        browser.quit()

def isNullFile(version, brand, hour) :
    result = openpyxl.load_workbook(wehagoReport(version, brand, hour))
    sheet = result.active
    if '테스트 결과 없음' in sheet['A1'].value :
        return True
    else :
        return False

class WehagoResult :
    # excel 열어서 활성화 시키기
    result = openpyxl.load_workbook(resultPath()+'/WEHAGO 결과 sample.xlsx')
    sheet = result.active

    #엑셀의 함수명 열 리스트로 받기
    cellList = []
    col = sheet['D']
    for cell in col:
        cellList.append(cell.value)

    def confirmRow (self, serviceName) :
        row = len(WehagoResult.cellList)
        # 함수명 확인하여 row 값 반환
        if serviceName in WehagoResult.cellList :
            #row가 몇번째인지 # 1을 더하는 이유는 0부터이기 때문, 
            row = WehagoResult.cellList.index(serviceName) + 1
        return row

    def testSuccess (self, serviceName) :
        #서비스명의 row가 몇번째인지 확인해서 그 아래 column 값을 pass로 변경
        row = self.confirmRow(serviceName)
        WehagoResult.sheet.cell(row,column=5, value='Pass')

    def testFailure (self, serviceName, exer, fileName) :
        row = self.confirmRow(serviceName)
        WehagoResult.sheet.cell(row,column=5, value='Fail')
        WehagoResult.sheet.cell(row,column=6, value=currentTime().strftime('%H:%M'))
        #fail된 사유 작성
        WehagoResult.sheet.cell(row,column=7,value=exer)
        #fail  되기 전 상황 캡쳐
        cellNum = 'H' + str(row)
        self.failScreenshot(fileName, cellNum)

    def failScreenshot (self, fileName, cellNum) :
        path = os.path.join(resultPath(), 'image')
        img = openpyxl.drawing.image.Image(path+'/'+fileName)
        # 사이즈 크기 조정
        img.width = img.width/2
        img.height = img.height/2
        WehagoResult.sheet.add_image(img, cellNum)

    def runTime(self, serviceName, start) :
        row = self.confirmRow(serviceName)
        sec = time.time()-start
        times = str(datetime.timedelta(seconds=sec)).split(".")
        times = times[0]
        WehagoResult.sheet.cell(row,column=9, value=times)
    
    def savexl (self, version) : 
        WehagoResult.result.save(detailPath()+testReport(version))
        print("save yy")

    def upload(self, browser, version, brand, hour=True) :
        browser.get('https://www.wehago.com/#/communication2/talk/VapzR3cBqq7xHM_q0h6e')
        time.sleep(5)
        if "자동화 리포트용" in browser.title :
            browser.find_element_by_css_selector("input[type='file']").send_keys(wehagoReport(version, brand, hour))
        time.sleep(3)

    def fileUnion(self, version, brand, hour=True) :
        #파일 Union  
        all_data = pd.DataFrame()
        noFail = pd.DataFrame({'실패건 없음!!'}) 
        noData = pd.DataFrame({'테스트 결과 없음'})

        # 저장할 파일
        result = pd.ExcelWriter(wehagoReport(version, brand, hour), engine='xlsxwriter')
        if hour :
            fileList = glob.glob(hourReport(version))
        else :
            fileList = glob.glob(dayReport(version))

        if fileList :
            # n일에 진행된 테스트 결과 취합
            for f in fileList: 
                df = pd.read_excel(f, usecols=[*range(1,7)])
                df = df[df['테스트 결과'] == 'Fail']
                all_data = all_data.append(df, ignore_index=True)
        
            # 테스트 결과 값에 Fail된 값이 있으면
            if not all_data[all_data['테스트 결과'] == 'Fail'].empty :
                #결과값 통계 생성
                all_data[all_data['테스트 결과'] == 'Fail'].groupby(['함수명', '테스트 결과']).size().sort_values(ascending=False).to_excel(result, sheet_name='결과')
            else :
                noFail.to_excel(result, sheet_name='결과',index=False, header=False)
            # 테스트 결과 취합하여 엑셀에 저장
            all_data.to_excel(result, sheet_name='통합',index=False)
            # #파일 저장
            result.save()
            self.adjustFiles(result)
        else :
            noData.to_excel(result, sheet_name='결과',index=False, header=False)
            result.save()
            print('통합 파일 없음')

    def adjustFiles(self, report) :
        result = openpyxl.load_workbook(report)
        sheet1 = result['결과']
        sheet1.column_dimensions['A'].width = 25
        sheet1 = result['통합']
        sheet1.column_dimensions['A'].width = 15
        sheet1.column_dimensions['B'].width = 25
        result.save(report)
